import openpyxl
from openpyxl.styles import Font, colors
from openpyxl.worksheet.hyperlink import Hyperlink


def update():
    wb = openpyxl.load_workbook('high.xlsx')

    sheetnames = wb.sheetnames
    print('sheetnames=', sheetnames)

    sheet = wb['Sheet1']

    # print('title=', sheet.title)

    # sheet = wb.active

    sheet['H8'] = 'good'

    # sheet['A10'] = 'You should see three logos below'
    # img = Image('img/a.png')
    # sheet.add_image(img, 'A10')

    h = Hyperlink('img/a.png', 'img/a.png')

    s = '=HYPERLINK("{}","截图")'.format('img/a.png')
    # sheet['H9'] = '=HYPERLINK("img/a.png","截图2")'
    sheet['H10'] = '=HYPERLINK("img/a.png","截图3")'
    sheet['H10'].font = Font(color=colors.BLUE, underline=Font.UNDERLINE_SINGLE)

    wb.save('high.xlsx')


update()
