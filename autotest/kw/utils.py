import time
import traceback

import openpyxl
from appium import webdriver
from openpyxl.styles import colors, Font
from selenium.webdriver.support.wait import WebDriverWait

from itcast.autotest.kw import constant
from itcast.autotest.kw.domain import CaseStep, PageElement, StepData, TestCase, CaseList


class DriverUtil:
    """
    驱动工具类
    """
    __driver = None

    @staticmethod
    def get_driver():
        print('get_driver...')
        if DriverUtil.__driver is None:
            DriverUtil.__init_driver()
        return DriverUtil.__driver

    @staticmethod
    def __init_driver():
        print('__init_driver...')
        cap = {'platformName': 'Android', 'deviceName': 'emulator', 'appPackage': 'net.csdn.csdnplus',
               'appActivity': '.activity.SplashActivity', 'automationName': 'Uiautomator2'}
        DriverUtil.__driver = webdriver.Remote('http://127.0.0.1:4723/wd/hub', cap)
        DriverUtil.__driver.implicitly_wait(10)

    @staticmethod
    def quit_driver():
        print('quit_driver...')
        DriverUtil.__driver.quit()


class ExcelUtil:
    """
    操作Excel的工具类
    """

    @staticmethod
    def set_cell_value(file_path, sheet_name, row, column, value):
        """
        设置单元格内容
        :param file_path: excel文件的路径
        :param sheet_name: sheet的名称
        :param row: 行
        :param column: 列
        :param value: 要设置的内容
        """
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]
        sheet.cell(row=row + 1, column=column + 1, value=value)
        wb.save(file_path)

    @staticmethod
    def set_cell_value_of_link(file_path, sheet_name, row, column, img_path):
        """
        向单元格中插入图片链接
        :param file_path: excel文件的路径
        :param sheet_name: sheet的名称
        :param row: 行
        :param column: 列
        :param img_path: 图片路径
        """
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]
        value = '=HYPERLINK("{}","截图")'.format(img_path)
        cell = sheet.cell(row=row + 1, column=column + 1, value=value)
        cell.font = Font(color=colors.BLUE, underline=Font.UNDERLINE_SINGLE)
        wb.save(file_path)

    @staticmethod
    def get_row_num(file_path, sheet_name, column, key_word):
        row_num = -1
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]
        for row in sheet.iter_rows():
            cell = row[column]
            if key_word == cell.value:
                row_num = cell.row
                break
        return row_num - 1

    @staticmethod
    def get_row_num2(file_path, sheet_name, column, key_word, column2, key_word2):
        row_num = -1
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]
        for row in sheet.iter_rows():
            cell = row[column]
            cell2 = row[column2]
            if key_word == cell.value and key_word2 == cell2.value:
                row_num = cell.row
                break
        return row_num - 1

    @staticmethod
    def get_all_data(file_path, sheet_name, start_row_num=0):
        """
        获取指定sheet中的全部数据
        :param file_path: excel文件的路径
        :param sheet_name: sheet的名称
        :param start_row_num: 指定开始行
        """
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]
        all_data = []
        for row in sheet.iter_rows(min_row=start_row_num + 1):
            row_data = []
            for cell in row:
                row_data.append(cell.value)
            all_data.append(row_data)
        return all_data


if __name__ == '__main__':
    # all_data = ExcelUtil.get_all_data('testcase/suite.xlsx', 'stepData', 1)
    # print(all_data)
    # ExcelUtil.set_cell_value_of_link('testcase/suite.xlsx', 'stepData', 10, 6, 'aaa.png')
    # ExcelUtil.set_cell_value_of_link('testcase/suite.xlsx', 'stepData', 11, 6, 'aaa.png')
    # ExcelUtil.set_cell_value_of_link('testcase/suite.xlsx', 'stepData', 15, 6, 'aaa.png')
    # ExcelUtil.set_cell_value('testcase/suite.xlsx', 'stepData', 16, 6, 'ttttttt')

    # wb = openpyxl.load_workbook('testcase/suite.xlsx')
    # sheet = wb['stepData']
    # # rows = sheet.rows
    # # print(tuple(rows))
    # all_data = []
    # for row in sheet.iter_rows(min_row=2):
    #     row_data = []
    #     for cell in row:
    #         row_data.append(cell.value)
    #         print(cell.row)
    #     all_data.append(row_data)
    # print(all_data)

    r = ExcelUtil.get_row_num('testcase/suite.xlsx', 'stepData', 2, 's_004')
    print('r=', r)


class TestCaseUtil:
    """
    用例工具类
    """

    @staticmethod
    def find_element(driver, case_step):
        """
        查找元素
        :param driver: 驱动
        :param case_step: 执行步骤
        :return: 该用例步骤操作的元素
        """
        element = None
        page_element = case_step.page_element
        location_type = page_element.location_type
        if 'id' == location_type:
            element = driver.find_element_by_id(page_element.location_value)
        elif 'xpath' == location_type:
            element = driver.find_element_by_xpath(page_element.location_value)
        elif 'toast' == location_type:
            xpath = ".//*[contains(@text,'{}')]".format(case_step.step_data.expected_result)
            element = WebDriverWait(driver, 10).until(lambda x: x.find_element_by_xpath(xpath))
        else:
            print('unknown location type=', location_type)
        return element

    @staticmethod
    def execute_operation(element, action, input_data, expected_result):
        """
        执行操作
        :param element: 元素
        :param action: 动作
        :param input_data: 输入数据
        :param expected_result: 期望结果
        :return: 是否操作成功[true:操作成功; false:操作失败]
        """
        if 'sendKey' == action:
            element.send_keys(input_data)
            return True
        elif 'click' == action:
            element.click()
            return True
        elif 'getText' == action:
            print('expected_result={} text={}'.format(expected_result, element.text))
            return element.text == expected_result
        elif 'getToast' == action:
            return element is not None
        else:
            print('unknown action=', action)
            return False

    @staticmethod
    def execute_case_step(driver, case_step):
        print('execute_case_step start execute case step={}...'.format(case_step.step_desc))
        is_success = False
        try:
            # 查找元素
            element = TestCaseUtil.find_element(driver, case_step)

            # 执行操作
            is_success = TestCaseUtil.execute_operation(element, case_step.action, case_step.step_data.input_data,
                                                        case_step.step_data.expected_result)
        except Exception:
            print('execute_case_step error=', traceback.format_exc())
        return is_success

    @staticmethod
    def execute_test_case(driver, test_case):
        all_step_success = True
        for case_step in test_case.case_step_list:
            is_success = TestCaseUtil.execute_case_step(driver, case_step)

            # 更新执行结果
            TestCaseUtil.update_execute_result(test_case, case_step, is_success)

            # 执行失败
            if not is_success:
                TestCaseUtil.save_screenshot(driver, test_case, case_step)

                all_step_success = False
                break

        if all_step_success:
            test_case.execute_result = constant.PASS
            TestCaseUtil.update_execute_result_of_test_case(test_case.case_code, constant.PASS)
        else:
            test_case.execute_result = constant.FAIL

        # 后置处理
        if test_case.post_process != '':
            if 'resetApp' == test_case.post_process:
                driver.reset()

    @staticmethod
    def save_screenshot(driver, test_case, case_step):
        print('save_screenshot case_code={} step_code={}'.format(test_case.case_code, case_step.step_code))
        file_name = "screenshot/" + str(time.time()) + '.png'
        driver.get_screenshot_as_file(constant.TEST_CASE_DATA_DIR + file_name)

        TestCaseUtil.update_screenshot_of_case_step(test_case.case_code, case_step.step_code, file_name)

    @staticmethod
    def update_execute_result(test_case, case_step, is_success):
        ret = constant.PASS
        if not is_success:
            ret = constant.FAIL
        case_step.step_execute_result = ret

        TestCaseUtil.update_execute_result_of_case_step(test_case.case_code, case_step.step_code, ret)

        if not is_success:
            TestCaseUtil.update_execute_result_of_test_case(test_case.case_code, ret)

    @staticmethod
    def update_execute_result_of_test_case(case_code, result):
        row_num = ExcelUtil.get_row_num(constant.SUITE_FILE_PATH, "testCase", 0, case_code)

        ExcelUtil.set_cell_value(constant.SUITE_FILE_PATH, "testCase", row_num, 3, result)

    @staticmethod
    def update_execute_result_of_case_step(case_code, step_code, result):
        row_num = ExcelUtil.get_row_num2(constant.SUITE_FILE_PATH, "stepData", 0, case_code, 1, step_code)

        ExcelUtil.set_cell_value(constant.SUITE_FILE_PATH, "stepData", row_num, 5, result)

    @staticmethod
    def update_screenshot_of_case_step(case_code, step_code, screenshot_path):
        print('update_screenshot_of_case_step...')
        row_num = ExcelUtil.get_row_num2(constant.SUITE_FILE_PATH, "stepData", 0, case_code, 1, step_code)

        print('update_screenshot_of_case_step...row_num=', row_num)
        ExcelUtil.set_cell_value_of_link(constant.SUITE_FILE_PATH, "stepData", row_num, 6, screenshot_path)

    @staticmethod
    def load_case_list():
        # 获取用例列表数据
        data_list = ExcelUtil.get_all_data(constant.SUITE_FILE_PATH, 'testCase', 1)

        # 组装数据
        case_list = CaseList()
        for data in data_list:
            test_case = TestCase(data[0], data[1], data[2])

            # 加载用例步骤数据
            TestCaseUtil.load_case_step(test_case)

            case_list.test_case_list.append(test_case)
        return case_list

    @staticmethod
    def load_case_step(test_case):
        # 获取用例步骤数据
        page = test_case.get_page()
        data_list = ExcelUtil.get_all_data(constant.TEST_CASE_DATA_DIR + page + ".xlsx", 'caseStep', 1)

        # 组装用例步骤数据
        for data in data_list:
            if test_case.case_code == data[0]:
                case_step = CaseStep(data[0], data[1], data[2], data[3], data[4])

                # 加载该步骤操作的元素信息
                page_element = TestCaseUtil.load_page_element(case_step, page)
                case_step.page_element = page_element

                # 加载该步骤的测试数据
                step_data = TestCaseUtil.load_step_data(test_case.case_code, case_step.step_code)
                case_step.step_data = step_data

                test_case.case_step_list.append(case_step)

    @staticmethod
    def load_step_data(case_code, step_code):
        # 获取步骤测试数据
        data_list = ExcelUtil.get_all_data(constant.SUITE_FILE_PATH, 'stepData', 1)

        # 组装数据
        for data in data_list:
            if case_code == data[0] and step_code == data[1]:
                return StepData(data[0], data[1], data[3], data[4])
        return None

    @staticmethod
    def load_page_element(case_step, page):
        # 获取页面元素信息数据
        data_list = ExcelUtil.get_all_data(constant.TEST_CASE_DATA_DIR + page + '.xlsx', 'location', 1)

        # 组装数据
        for data in data_list:
            if case_step.element_name == data[0]:
                return PageElement(data[0], data[1], data[2], data[3])
        return None


class ReportUtil:
    """
    测试报告工具类
    """

    @staticmethod
    def create_report(case_list, report):
        """
        生成测试报告
        :param case_list: 用例列表
        :param report: 报告数据
        """
        report.totalTests = len(case_list.test_case_list)

        total_steps = 0
        test_pass_count = 0
        test_fail_count = 0
        test_skip_count = 0
        step_pass_count = 0
        step_fail_count = 0
        step_skip_count = 0

        for test_case in case_list.test_case_list:
            if constant.PASS == test_case.execute_result:
                test_pass_count += 1
            elif constant.FAIL == test_case.execute_result:
                test_fail_count += 1
            else:
                test_skip_count += 1

            for case_step in test_case.case_step_list:
                total_steps += 1
                if constant.PASS == case_step.step_execute_result:
                    step_pass_count += 1
                elif constant.FAIL == case_step.step_execute_result:
                    step_fail_count += 1
                else:
                    step_skip_count += 1

        report.totalSteps = total_steps
        report.testPassCount = test_pass_count
        report.testFailCount = test_fail_count
        report.testSkipCount = test_skip_count
        report.stepPassCount = step_pass_count
        report.stepFailCount = step_fail_count
        report.stepSkipCount = step_skip_count

        ReportUtil.update_report_to_excel(report)

    @staticmethod
    def update_report_to_excel(report):
        wb = openpyxl.load_workbook(constant.SUITE_FILE_PATH)
        sheet = wb['report']

        # 概述
        sheet.cell(row=3, column=1, value=report.totalTests)
        sheet.cell(row=3, column=2, value=report.totalSteps)
        sheet.cell(row=3, column=3, value=report.takeTime)

        # 用例统计
        sheet.cell(row=7, column=1, value=report.testPassCount)
        sheet.cell(row=7, column=2, value=report.testFailCount)
        sheet.cell(row=7, column=3, value=report.testSkipCount)

        # 步骤统计
        sheet.cell(row=11, column=1, value=report.stepPassCount)
        sheet.cell(row=11, column=2, value=report.stepFailCount)
        sheet.cell(row=11, column=3, value=report.stepSkipCount)

        wb.save(constant.SUITE_FILE_PATH)
